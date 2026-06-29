from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import DAILY_LOGS_DIR, HISTORY_DIR, FIXED_WORKERS
from services.database import get_db
from services.database import (
    get_daily_totals,
    get_active_workers,
    get_all_products,
    get_worker_month_production,
    get_worker_daily_breakdown,
)


def _get_date_absence_info(date_str: str) -> tuple[bool, str | None]:
    conn = get_db()
    worker_count = len(FIXED_WORKERS)
    absent = conn.execute(
        """SELECT COUNT(*) AS cnt, reason FROM daily_log
           WHERE entry_date = ? AND status = 'absent'
           GROUP BY reason ORDER BY cnt DESC LIMIT 1""",
        (date_str,),
    ).fetchone()
    if absent and absent["cnt"] >= worker_count:
        return (True, absent["reason"] or None)
    return (False, None)


def generate_excel_report(period: str = "daily", year: Optional[int] = None, month: Optional[int] = None, day: Optional[int] = None) -> str:
    today = date.today()
    y = year or today.year
    m = month or today.month
    d = day or today.day

    products = get_all_products()
    product_codes = [p["code"] for p in products]

    if period == "daily":
        return _daily_excel_report(y, m, d, product_codes)
    elif period == "weekly":
        return _weekly_excel_report(y, m, d, product_codes)
    elif period == "monthly":
        return _monthly_excel_report(y, m, product_codes)
    return f"Unknown period '{period}'"


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(ws, row: int, cols: int):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = _thin_border()


def _daily_excel_report(year: int, month: int, day: int, product_codes: list[str]) -> str:
    date_str = f"{year}-{month:02d}-{day:02d}"
    totals = get_daily_totals(date_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Daily {day:02d}-{month:02d}"

    ws.cell(row=1, column=1, value=f"Daily Production Report - {date_str}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(product_codes) + 1)

    headers = ["Product"] + product_codes
    _write_header_row(ws, 3, headers)

    row = 4
    for code in product_codes:
        qty = totals.get(code, 0)
        ws.cell(row=row, column=1, value=code).border = _thin_border()
        ws.cell(row=row, column=2, value=qty).border = _thin_border()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    filepath = DAILY_LOGS_DIR / f"report_daily_{date_str}.xlsx"
    wb.save(filepath)
    return str(filepath)


def _weekly_excel_report(year: int, month: int, day: int, product_codes: list[str]) -> str:
    from datetime import date as dt_date, timedelta
    dt = dt_date(year, month, day)
    monday = dt - timedelta(days=dt.weekday())
    sunday = monday + timedelta(days=6)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Weekly {monday.isoformat()}"

    ws.cell(row=1, column=1, value=f"Weekly Production Report - {monday.isoformat()} to {sunday.isoformat()}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(product_codes) + 2)

    headers = ["Date", "Day"] + product_codes
    _write_header_row(ws, 3, headers)

    row = 4
    for i in range(7):
        d = monday + timedelta(days=i)
        ds = d.isoformat()
        totals = get_daily_totals(ds)
        ws.cell(row=row, column=1, value=ds).border = _thin_border()
        ws.cell(row=row, column=2, value=d.strftime("%A")).border = _thin_border()
        for j, code in enumerate(product_codes):
            qty = totals.get(code, 0)
            cell = ws.cell(row=row, column=3 + j, value=qty)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(67 + j)].width = 12

    filepath = DAILY_LOGS_DIR / f"report_weekly_{monday.isoformat()}.xlsx"
    wb.save(filepath)
    return str(filepath)


def _monthly_excel_report(year: int, month: int, product_codes: list[str]) -> str:
    from calendar import monthrange
    days = monthrange(year, month)[1]
    num_cols = len(product_codes) + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly {year}-{month:02d}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    name_cell = ws.cell(row=1, column=1, value="MONTHLY REPORT")
    name_cell.font = Font(name="Calibri", bold=True, italic=True, size=18)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.border = Border()
    ws.row_dimensions[1].height = 28
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).border = Border()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    month_cell = ws.cell(row=2, column=1, value=f"{year}-{month:02d}")
    month_cell.font = Font(name="Calibri", size=11, color="666666")
    month_cell.alignment = Alignment(horizontal="center")
    month_cell.border = Border()
    for c in range(2, num_cols + 1):
        ws.cell(row=2, column=c).border = Border()
    ws.row_dimensions[2].height = 18

    headers = ["Date"] + product_codes
    _write_header_row(ws, 4, headers)

    row = 5
    grand_totals = {code: 0 for code in product_codes}
    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    absent_font = Font(name="Calibri", color="D93025", bold=True)
    for d in range(1, days + 1):
        date_str = f"{year}-{month:02d}-{d:02d}"
        all_absent, reason = _get_date_absence_info(date_str)
        totals = get_daily_totals(date_str)
        cell = ws.cell(row=row, column=1, value=date_str)
        cell.border = _thin_border()
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center")
        if all_absent and reason:
            cell.fill = absent_fill
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
            rc = ws.cell(row=row, column=2, value=reason)
            rc.border = _thin_border()
            rc.fill = absent_fill
            rc.font = absent_font
            rc.alignment = Alignment(horizontal="center")
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = _thin_border()
                ws.cell(row=row, column=c).fill = absent_fill
        else:
            for j, code in enumerate(product_codes):
                qty = totals.get(code, 0)
                cell = ws.cell(row=row, column=2 + j, value=qty)
                cell.border = _thin_border()
                cell.font = Font(name="Calibri")
                cell.alignment = Alignment(horizontal="center")
                grand_totals[code] += qty
        row += 1

    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(name="Calibri", bold=True, size=11)
    total_cell.border = _thin_border()
    total_cell.alignment = Alignment(horizontal="center")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for j, code in enumerate(product_codes):
        cell = ws.cell(row=row, column=2 + j, value=grand_totals[code])
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = total_fill

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 14

    filepath = DAILY_LOGS_DIR / f"report_monthly_{year}-{month:02d}.xlsx"
    wb.save(filepath)
    return str(filepath)


def generate_worker_excel(worker_name: str, year: int, month: int) -> Optional[str]:
    from calendar import monthrange
    from services.database import get_worker_id

    wid = get_worker_id(worker_name)
    if not wid:
        return None

    products = get_all_products()
    product_codes = [p["code"] for p in products]
    days_data = get_worker_daily_breakdown(wid, year, month)

    if not days_data:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{worker_name} {year}-{month:02d}"

    num_cols = len(product_codes) + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    name_cell = ws.cell(row=1, column=1, value=worker_name)
    name_cell.font = Font(name="Calibri", bold=True, italic=True, size=18)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.border = Border()
    ws.row_dimensions[1].height = 28
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).border = Border()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    month_cell = ws.cell(row=2, column=1, value=f"{year}-{month:02d}")
    month_cell.font = Font(name="Calibri", size=11, color="666666")
    month_cell.alignment = Alignment(horizontal="center")
    month_cell.border = Border()
    for c in range(2, num_cols + 1):
        ws.cell(row=2, column=c).border = Border()
    ws.row_dimensions[2].height = 18

    headers = ["Date"] + product_codes
    _write_header_row(ws, 4, headers)

    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    absent_font = Font(name="Calibri", color="D93025", bold=True)

    row = 5
    grand_totals = {code: 0 for code in product_codes}
    date_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    for day in days_data:
        date_cell = ws.cell(row=row, column=1, value=day["date"])
        date_cell.border = _thin_border()
        date_cell.font = Font(name="Calibri", bold=True)
        date_cell.alignment = Alignment(horizontal="center")

        if day["status"] == "absent":
            reason_text = day.get("reason") or "ABSENT"
            cell = ws.cell(row=row, column=2, value=reason_text)
            cell.border = _thin_border()
            cell.fill = absent_fill
            cell.font = absent_font
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = _thin_border()
                ws.cell(row=row, column=c).fill = absent_fill
            row += 1
            continue

        for j, code in enumerate(product_codes):
            qty = day["products"].get(code, 0)
            cell = ws.cell(row=row, column=2 + j, value=qty)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
            grand_totals[code] += qty
        row += 1

    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(name="Calibri", bold=True, size=11)
    total_cell.border = _thin_border()
    total_cell.alignment = Alignment(horizontal="center")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for j, code in enumerate(product_codes):
        cell = ws.cell(row=row, column=2 + j, value=grand_totals[code])
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = total_fill

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 14

    filepath = HISTORY_DIR / f"{worker_name}_{year}_{month:02d}.xlsx"
    wb.save(filepath)
    return str(filepath)


def generate_worker_excel_stream(worker_name: str, year: int, month: int) -> Optional[tuple[BytesIO, str]]:
    from calendar import monthrange
    from services.database import get_worker_id

    wid = get_worker_id(worker_name)
    if not wid:
        return None

    products = get_all_products()
    product_codes = [p["code"] for p in products]
    days_data = get_worker_daily_breakdown(wid, year, month)

    if not days_data:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{worker_name} {year}-{month:02d}"

    num_cols = len(product_codes) + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    name_cell = ws.cell(row=1, column=1, value=worker_name)
    name_cell.font = Font(name="Calibri", bold=True, italic=True, size=18)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.border = Border()
    ws.row_dimensions[1].height = 28
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).border = Border()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    month_cell = ws.cell(row=2, column=1, value=f"{year}-{month:02d}")
    month_cell.font = Font(name="Calibri", size=11, color="666666")
    month_cell.alignment = Alignment(horizontal="center")
    month_cell.border = Border()
    for c in range(2, num_cols + 1):
        ws.cell(row=2, column=c).border = Border()
    ws.row_dimensions[2].height = 18

    headers = ["Date"] + product_codes
    _write_header_row(ws, 4, headers)

    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    absent_font = Font(name="Calibri", color="D93025", bold=True)

    row = 5
    grand_totals = {code: 0 for code in product_codes}

    for day in days_data:
        date_cell = ws.cell(row=row, column=1, value=day["date"])
        date_cell.border = _thin_border()
        date_cell.font = Font(name="Calibri", bold=True)
        date_cell.alignment = Alignment(horizontal="center")

        if day["status"] == "absent":
            reason_text = day.get("reason") or "ABSENT"
            cell = ws.cell(row=row, column=2, value=reason_text)
            cell.border = _thin_border()
            cell.fill = absent_fill
            cell.font = absent_font
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = _thin_border()
                ws.cell(row=row, column=c).fill = absent_fill
            row += 1
            continue

        for j, code in enumerate(product_codes):
            qty = day["products"].get(code, 0)
            cell = ws.cell(row=row, column=2 + j, value=qty)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
            grand_totals[code] += qty
        row += 1

    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(name="Calibri", bold=True, size=11)
    total_cell.border = _thin_border()
    total_cell.alignment = Alignment(horizontal="center")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for j, code in enumerate(product_codes):
        cell = ws.cell(row=row, column=2 + j, value=grand_totals[code])
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = total_fill

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{worker_name}_{year}_{month:02d}.xlsx"
    return (buf, filename)


def generate_monthly_excel_stream(year: int, month: int) -> tuple[BytesIO, str]:
    from calendar import monthrange
    products = get_all_products()
    product_codes = [p["code"] for p in products]
    days = monthrange(year, month)[1]
    num_cols = len(product_codes) + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly {year}-{month:02d}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    name_cell = ws.cell(row=1, column=1, value="MONTHLY REPORT")
    name_cell.font = Font(name="Calibri", bold=True, italic=True, size=18)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.border = Border()
    ws.row_dimensions[1].height = 28
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).border = Border()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    month_cell = ws.cell(row=2, column=1, value=f"{year}-{month:02d}")
    month_cell.font = Font(name="Calibri", size=11, color="666666")
    month_cell.alignment = Alignment(horizontal="center")
    month_cell.border = Border()
    for c in range(2, num_cols + 1):
        ws.cell(row=2, column=c).border = Border()
    ws.row_dimensions[2].height = 18

    headers = ["Date"] + product_codes
    _write_header_row(ws, 4, headers)

    row = 5
    grand_totals = {code: 0 for code in product_codes}
    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    absent_font = Font(name="Calibri", color="D93025", bold=True)
    for d in range(1, days + 1):
        date_str = f"{year}-{month:02d}-{d:02d}"
        all_absent, reason = _get_date_absence_info(date_str)
        totals = get_daily_totals(date_str)
        cell = ws.cell(row=row, column=1, value=date_str)
        cell.border = _thin_border()
        cell.font = Font(name="Calibri", bold=True)
        cell.alignment = Alignment(horizontal="center")
        if all_absent and reason:
            cell.fill = absent_fill
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
            rc = ws.cell(row=row, column=2, value=reason)
            rc.border = _thin_border()
            rc.fill = absent_fill
            rc.font = absent_font
            rc.alignment = Alignment(horizontal="center")
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = _thin_border()
                ws.cell(row=row, column=c).fill = absent_fill
        else:
            for j, code in enumerate(product_codes):
                qty = totals.get(code, 0)
                cell = ws.cell(row=row, column=2 + j, value=qty)
                cell.border = _thin_border()
                cell.font = Font(name="Calibri")
                cell.alignment = Alignment(horizontal="center")
                grand_totals[code] += qty
        row += 1

    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(name="Calibri", bold=True, size=11)
    total_cell.border = _thin_border()
    total_cell.alignment = Alignment(horizontal="center")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for j, code in enumerate(product_codes):
        cell = ws.cell(row=row, column=2 + j, value=grand_totals[code])
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = total_fill

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"report_monthly_{year}-{month:02d}.xlsx"
    return (buf, filename)


def _write_header_row(ws, row: int, headers: list[str]):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = _thin_border()
