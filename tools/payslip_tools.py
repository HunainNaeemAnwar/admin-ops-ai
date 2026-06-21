from datetime import date

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from config import PDF_DIR, EXCEL_SLIPS_DIR, FIXED_WORKERS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _format_month(month: int) -> str:
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    return months[month - 1]


def _build_payslip_data(worker: str, year: int, month: int) -> dict | None:
    from tools.database import get_worker_id, get_worker_month_production, get_payslip
    from tools.database import get_all_products
    from tools.rejection_tools import get_distribution_for_month
    from config import TAX_PERCENTAGE

    wid = get_worker_id(worker)
    if not wid:
        return None

    ps = get_payslip(wid, year, month)
    if ps:
        return dict(ps)

    production = get_worker_month_production(wid, year, month)
    if not production:
        return None

    products = get_all_products()
    product_rates = {p["code"]: p["rate"] for p in products}

    product_totals = {}
    for p in production:
        code = p["product_code"]
        product_totals[code] = product_totals.get(code, 0) + p["quantity"]

    gross_total = sum(qty * product_rates.get(code, 0) for code, qty in product_totals.items())

    distribution = get_distribution_for_month(year, month)
    rejection_deduction = 0
    for dist in distribution:
        w_share = dist["distribution"].get(worker, 0)
        rejection_deduction += w_share * product_rates.get(dist["product_code"], 0)

    from tools.database import get_total_advances_for_worker_month
    advance_deduction = get_total_advances_for_worker_month(wid, year, month)

    tax_amount = round(gross_total * TAX_PERCENTAGE / 100, 2)
    net_payable = round(gross_total - rejection_deduction - advance_deduction - tax_amount, 2)

    return {
        "gross_total": gross_total,
        "tax_total": tax_amount,
        "rejection_deduction": rejection_deduction,
        "advance_deduction": advance_deduction,
        "net_payable": net_payable,
        "product_totals": product_totals,
    }


def generate_pdf_payslip(worker: str, year: int, month: int) -> str:
    data = _build_payslip_data(worker, year, month)
    if not data:
        return f"No data for {worker} in {year}-{month:02d}"

    filename = f"{worker}_{year}_{month:02d}.pdf"
    filepath = PDF_DIR / filename
    doc = SimpleDocTemplate(
        str(filepath), pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], spaceAfter=6*mm)
    normal = styles["Normal"]
    elements = []

    elements.append(Paragraph(f"Pay Slip - {_format_month(month)} {year}", title_style))
    elements.append(Paragraph(f"Worker: {worker}", normal))
    elements.append(Spacer(1, 6*mm))

    products = data.get("product_totals", {})
    prod_data = [[
        Paragraph("<b>Product</b>", normal),
        Paragraph("<b>Qty</b>", normal),
        Paragraph("<b>Gross (Rs)</b>", normal),
    ]]
    for code, qty in sorted(products.items()):
        prod_data.append([code, str(qty), ""])
    total_line = [Paragraph("<b>Total</b>", normal), "", f'{data["gross_total"]:,.2f}']
    prod_data.append(total_line)

    col_widths = [50*mm, 30*mm, 40*mm]
    t = Table(prod_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6*mm))

    summary_data = [
        ["Description", "Amount (Rs)"],
        ["Gross Total", f'{data["gross_total"]:,.2f}'],
        [f"Rejection Deduction", f'({data["rejection_deduction"]:,.2f})'],
        [f"Advance Deduction", f'({data["advance_deduction"]:,.2f})'],
        [f"Tax Deducted (3%)", f'({data["tax_total"]:,.2f})'],
        ["Net Payable", f'{data["net_payable"]:,.2f}'],
    ]
    st = Table(summary_data, colWidths=[80*mm, 50*mm])
    st.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(f"Generated on: {date.today().isoformat()}", normal))

    doc.build(elements)
    return str(filepath)


def generate_excel_payslip(worker: str, year: int, month: int) -> str:
    data = _build_payslip_data(worker, year, month)
    if not data:
        return f"No data for {worker} in {year}-{month:02d}"

    filename = f"{worker}_{year}_{month:02d}.xlsx"
    filepath = EXCEL_SLIPS_DIR / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pay Slip"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws.cell(row=1, column=1, value=f"Pay Slip - {_format_month(month)} {year}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Worker: {worker}").font = Font(bold=True, size=12)
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    headers = ["Item", "Qty", "Amount (Rs)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 5
    products = data.get("product_totals", {})
    for code, qty in sorted(products.items()):
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=qty)
        row += 1

    ws.cell(row=row, column=1, value="Gross Total").font = Font(bold=True)
    ws.cell(row=row, column=3, value=data["gross_total"]).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Rejection Deduction").font = Font(bold=True)
    ws.cell(row=row, column=3, value=data["rejection_deduction"]).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Advance Deduction").font = Font(bold=True)
    ws.cell(row=row, column=3, value=data["advance_deduction"]).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Tax Deducted").font = Font(bold=True)
    ws.cell(row=row, column=3, value=data["tax_total"]).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Net Payable").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=data["net_payable"]).font = Font(bold=True, size=12, color="006400")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    wb.save(filepath)
    return str(filepath)
