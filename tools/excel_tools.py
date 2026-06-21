import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from calendar import monthrange

from config import DAILY_LOGS_DIR, PRODUCT_CATALOG_PATH

DATA_SHEET = "_data"
TOTAL_SHEET = "DEPART TOTAL"
DATA_HEADERS = [
    "entry_id", "date", "worker", "product_code", "description",
    "quantity", "rate", "gross", "tax_pct", "tax_amt", "net", "timestamp"
]

TEMPLATE_COLS = ["DATE", "NUT", "10*20", "6*25", "6*30", "10*25"]
PRODUCT_COLUMN_MAP = {
    "NUT": ["NUT-STD", "NUT-M10"],
    "10*20": ["BOLT-10x20"],
    "6*25": ["BOLT-6x25"],
    "6*30": ["BOLT-6x30"],
    "10*25": ["BOLT-10x25"],
}

FIXED_WORKERS = ["Naeem", "Kaleem", "Akbar", "Suny", "Sajjad", "Irfan", "Kashif", "Gulmast"]

COL_COUNT = len(TEMPLATE_COLS)
MERGE_RANGE = f"A1:{chr(64 + COL_COUNT)}9"


def _get_monthly_path(year: int, month: int) -> Path:
    return DAILY_LOGS_DIR / f"{year}-{month:02d}.xlsx"


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _product_to_template_col(product_code: str) -> str | None:
    code = product_code.strip().upper()
    for col_name, codes in PRODUCT_COLUMN_MAP.items():
        if code in [c.upper() for c in codes]:
            return col_name
    return None


def _col_letter(c: int) -> str:
    return chr(64 + c)


def create_monthly_file(year: int, month: int):
    path = _get_monthly_path(year, month)
    wb = openpyxl.Workbook()

    ws_data = wb.active
    ws_data.title = DATA_SHEET
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(DATA_HEADERS, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    col_widths = [10, 14, 20, 18, 25, 10, 10, 12, 8, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws_data.column_dimensions[_col_letter(i)].width = w

    ws_total = wb.create_sheet(TOTAL_SHEET)
    _build_total_sheet(ws_total)

    for worker in FIXED_WORKERS:
        _build_worker_sheet(wb, worker, year, month)

    wb.save(path)
    return wb


def _build_total_sheet(ws):
    thin = _thin_border()
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Worker"] + TEMPLATE_COLS[1:] + ["Total Pieces"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = thin

    ws.column_dimensions["A"].width = 20
    for c_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[c_letter].width = 14


def _build_worker_sheet(wb, worker: str, year: int, month: int):
    days = monthrange(year, month)[1]
    ws = wb.create_sheet(title=worker.strip().title())
    thin = _thin_border()
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(MERGE_RANGE)
    name_cell = ws["A1"]
    name_cell.value = worker.upper()
    name_cell.font = Font(bold=True, size=14)
    name_cell.alignment = center
    for r in range(1, 10):
        for c in range(1, COL_COUNT + 1):
            ws.cell(row=r, column=c).border = thin

    for c, h in enumerate(TEMPLATE_COLS, 1):
        cell = ws.cell(row=10, column=c, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = thin

    for d in range(1, days + 1):
        row_num = 10 + d
        date_val = date(year, month, d)
        cell = ws.cell(row=row_num, column=1, value=date_val)
        cell.number_format = "DD-MM-YYYY"
        cell.border = thin
        for c in range(2, COL_COUNT + 1):
            ws.cell(row=row_num, column=c).value = 0
            ws.cell(row=row_num, column=c).border = thin
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

    total_row = 10 + days + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).border = thin
    for c in range(2, COL_COUNT + 1):
        first_data = 11
        last_data = 10 + days
        cell = ws.cell(row=total_row, column=c)
        cell.value = f"=SUM({_col_letter(c)}{first_data}:{_col_letter(c)}{last_data})"
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    reject_row = total_row + 2
    ws.cell(row=reject_row, column=1, value="REJECT").font = Font(bold=True, color="FF0000")
    ws.cell(row=reject_row, column=1).alignment = center
    ws.cell(row=reject_row, column=1).border = thin

    net_row = reject_row + 2
    ws.cell(row=net_row, column=1, value="NET").font = bold
    ws.cell(row=net_row, column=1).alignment = center
    ws.cell(row=net_row, column=1).border = thin
    for c in range(2, COL_COUNT + 1):
        total_cell = ws.cell(row=total_row, column=c)
        reject_cell = ws.cell(row=reject_row, column=c)
        net_cell = ws.cell(row=net_row, column=c)
        total_ref = f"{_col_letter(c)}{total_row}"
        reject_ref = f"{_col_letter(c)}{reject_row}"
        net_cell.value = f"={total_ref}-{reject_ref}"
        net_cell.font = bold
        net_cell.alignment = Alignment(horizontal="center")
        net_cell.border = thin

    for c in range(2, COL_COUNT + 1):
        cell = ws.cell(row=reject_row, column=c)
        cell.value = 0
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(color="FF0000")

    ws.column_dimensions["A"].width = 14
    for c_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[c_letter].width = 12

    return ws


def _get_or_create_worker_sheet(wb, worker: str, year: int, month: int):
    worker_safe = worker.strip().title()
    for s in wb.sheetnames:
        if s.lower() == worker_safe.lower():
            return wb[s]
    return _build_worker_sheet(wb, worker_safe, year, month)


def append_work_entry(
    worker: str,
    product_code: str,
    description: str,
    quantity: int,
    rate: float,
    gross: float,
    tax_pct: float,
    tax_amt: float,
    net: float,
    entry_date: Optional[str] = None,
):
    if entry_date is None:
        entry_date = date.today().isoformat()
    dt = date.fromisoformat(entry_date)
    year, month = dt.year, dt.month
    path = _get_monthly_path(year, month)

    if not path.exists():
        wb = create_monthly_file(year, month)
    else:
        wb = openpyxl.load_workbook(path)

    ws_data = wb[DATA_SHEET]
    if ws_data.max_row > 1:
        last_id = ws_data.cell(row=ws_data.max_row, column=1).value
        entry_id = (last_id or 0) + 1
    else:
        entry_id = 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_data.append([
        entry_id, entry_date, worker, product_code, description,
        quantity, rate, gross, tax_pct, tax_amt, net, now,
    ])

    col_widths = [10, 14, 20, 18, 25, 10, 10, 12, 8, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws_data.column_dimensions[_col_letter(i)].width = w

    ws_worker = _get_or_create_worker_sheet(wb, worker, year, month)
    day = dt.day
    row_num = 10 + day
    col_name = _product_to_template_col(product_code)
    if col_name:
        col_idx = TEMPLATE_COLS.index(col_name) + 1
        current = ws_worker.cell(row=row_num, column=col_idx).value or 0
        ws_worker.cell(row=row_num, column=col_idx).value = current + quantity

    _update_total_sheet(wb, year, month)
    wb.save(path)


def record_rejection(
    worker: str,
    product_code: str,
    description: str,
    quantity: int,
    rate: float,
    gross: float,
    tax_pct: float,
    tax_amt: float,
    net: float,
    entry_date: Optional[str] = None,
):
    if entry_date is None:
        entry_date = date.today().isoformat()
    dt = date.fromisoformat(entry_date)
    year, month = dt.year, dt.month
    path = _get_monthly_path(year, month)
    if not path.exists():
        create_monthly_file(year, month)
    wb = openpyxl.load_workbook(path)

    ws_data = wb[DATA_SHEET]
    if ws_data.max_row > 1:
        last_id = ws_data.cell(row=ws_data.max_row, column=1).value
        entry_id = (last_id or 0) + 1
    else:
        entry_id = 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    neg_qty = -abs(quantity)
    neg_gross = -abs(gross)
    neg_net = -abs(net)
    ws_data.append([
        entry_id, entry_date, worker, product_code,
        f"REJECT: {description}",
        neg_qty, rate, neg_gross, tax_pct, tax_amt, neg_net, now,
    ])

    ws_worker = _get_or_create_worker_sheet(wb, worker, year, month)
    days = monthrange(year, month)[1]
    reject_row = 10 + days + 4
    col_name = _product_to_template_col(product_code)
    if col_name:
        col_idx = TEMPLATE_COLS.index(col_name) + 1
        current = ws_worker.cell(row=reject_row, column=col_idx).value or 0
        ws_worker.cell(row=reject_row, column=col_idx).value = current + abs(quantity)

    _update_total_sheet(wb, year, month)
    wb.save(path)


def _update_total_sheet(wb, year: int, month: int):
    if TOTAL_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TOTAL_SHEET)
        _build_total_sheet(ws)
    else:
        ws = wb[TOTAL_SHEET]

    total_cols = COL_COUNT  # product columns (without DATE)
    total_width = total_cols + 2  # Worker + products + Total Pieces

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=total_width):
        for cell in row:
            cell.value = None

    worker_sheets = [s for s in wb.sheetnames if s not in (DATA_SHEET, TOTAL_SHEET)]
    thin = _thin_border()
    center = Alignment(horizontal="center", vertical="center")

    for i, w_name in enumerate(sorted(worker_sheets)):
        row_num = i + 2
        days = monthrange(year, month)[1]
        net_row = 10 + days + 6
        sheet_ref = f"'{w_name}'"

        ws.cell(row=row_num, column=1, value=w_name).border = thin
        for c in range(2, total_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.value = f"={sheet_ref}!{_col_letter(c)}{net_row}"
            cell.border = thin
            cell.alignment = center

        total_pieces_col = total_cols + 1
        cell = ws.cell(row=row_num, column=total_pieces_col)
        pieces_parts = "+".join(f"{_col_letter(c)}{row_num}" for c in range(2, total_cols + 1))
        cell.value = f"={pieces_parts}"
        cell.border = thin
        cell.alignment = center

    grand_row = len(worker_sheets) + 2
    ws.cell(row=grand_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=grand_row, column=1).border = thin
    for c in range(2, total_width + 1):
        cell = ws.cell(row=grand_row, column=c)
        if worker_sheets:
            first_data = 2
            last_data = grand_row - 1
            cell.value = f"=SUM({_col_letter(c)}{first_data}:{_col_letter(c)}{last_data})"
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = thin


def read_month_entries(year: int, month: int) -> list[dict]:
    path = _get_monthly_path(year, month)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb[DATA_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        entries.append({
            "entry_id": row[0],
            "date": str(row[1]) if row[1] else "",
            "worker": row[2] or "",
            "product_code": row[3] or "",
            "description": row[4] or "",
            "quantity": row[5] or 0,
            "rate": row[6] or 0.0,
            "gross": row[7] or 0.0,
            "tax_pct": row[8] or 0.0,
            "tax_amt": row[9] or 0.0,
            "net": row[10] or 0.0,
            "timestamp": str(row[11]) if row[11] else "",
        })
    return entries


def get_daily_total(year: int, month: int, day: int) -> float:
    entries = read_month_entries(year, month)
    date_str = f"{year}-{month:02d}-{day:02d}"
    return sum(e["net"] for e in entries if e["date"] == date_str)


def get_monthly_total(year: int, month: int) -> float:
    entries = read_month_entries(year, month)
    return sum(e["net"] for e in entries)


def get_worker_monthly_total(worker: str, year: int, month: int) -> float:
    entries = read_month_entries(year, month)
    return sum(e["net"] for e in entries if e["worker"].lower() == worker.lower())


def get_worker_entries(worker: str, year: int, month: int) -> list[dict]:
    entries = read_month_entries(year, month)
    return [e for e in entries if e["worker"].lower() == worker.lower()]


def get_all_workers(year: int, month: int) -> list[str]:
    return list(FIXED_WORKERS)


def has_today_entries(year: int, month: int, day: int) -> bool:
    entries = read_month_entries(year, month)
    date_str = f"{year}-{month:02d}-{day:02d}"
    return any(e["date"] == date_str for e in entries)


def load_product_catalog() -> list[dict]:
    if not PRODUCT_CATALOG_PATH.exists():
        return []
    wb = openpyxl.load_workbook(PRODUCT_CATALOG_PATH)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        products.append({
            "product_code": str(row[0]).strip(),
            "description": str(row[1] or "").strip(),
            "rate_per_piece": float(row[2] or 0),
            "tax_pct": float(row[3] or 0),
            "effective_from": str(row[4] or ""),
            "effective_to": str(row[5] or ""),
        })
    return products


def lookup_product(product_code: str) -> dict | None:
    products = load_product_catalog()
    code_lower = product_code.strip().lower()
    for p in products:
        if p["product_code"].lower() == code_lower:
            return p
    for p in products:
        if code_lower in p["product_code"].lower():
            return p
    return None


def update_product_rate(product_code: str, new_rate: float) -> dict:
    if not PRODUCT_CATALOG_PATH.exists():
        return {"error": "Product catalog not found"}
    wb = openpyxl.load_workbook(PRODUCT_CATALOG_PATH)
    ws = wb.active
    code_lower = product_code.strip().lower()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == code_lower:
            old_rate = row[2].value
            row[2].value = new_rate
            wb.save(PRODUCT_CATALOG_PATH)
            return {
                "product_code": str(cell.value).strip(),
                "description": str(row[1].value or "").strip(),
                "old_rate": old_rate,
                "new_rate": new_rate,
            }
    return {"error": f"Product '{product_code}' not found in catalog"}
