from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import DAILY_LOGS_DIR, HISTORY_DIR, FIXED_WORKERS
from tools.database import (
    get_daily_totals,
    get_active_workers,
    get_all_products,
    get_worker_month_production,
    get_worker_daily_breakdown,
)


def generate_excel_report(period: str = "daily", year: Optional[int] = None, month: Optional[int] = None, day: Optional[int] = None) -> str:
    today = date.today()
    y = year or today.year
    m = month or today.month
    d = day or today.day

    products = get_all_products()
    product_codes = [p["code"] for p in products]

    if period == "daily":
        return _daily_excel_report(y, m, d, product_codes)
    elif period == "weekly":
        return _weekly_excel_report(y, m, d, product_codes)
    elif period == "monthly":
        return _monthly_excel_report(y, m, product_codes)
    return f"Unknown period '{period}'"


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(ws, row: int, cols: int):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = _thin_border()


def _daily_excel_report(year: int, month: int, day: int, product_codes: list[str]) -> str:
    date_str = f"{year}-{month:02d}-{day:02d}"
    totals = get_daily_totals(date_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Daily {day:02d}-{month:02d}"

    ws.cell(row=1, column=1, value=f"Daily Production Report - {date_str}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(product_codes) + 1)

    headers = ["Product"] + product_codes
    _style_header(ws, 3, len(headers))

    row = 4
    for code in product_codes:
        qty = totals.get(code, 0)
        ws.cell(row=row, column=1, value=code).border = _thin_border()
        ws.cell(row=row, column=2, value=qty).border = _thin_border()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    filepath = DAILY_LOGS_DIR / f"report_daily_{date_str}.xlsx"
    wb.save(filepath)
    return str(filepath)


def _weekly_excel_report(year: int, month: int, day: int, product_codes: list[str]) -> str:
    from datetime import date as dt_date, timedelta
    dt = dt_date(year, month, day)
    monday = dt - timedelta(days=dt.weekday())
    sunday = monday + timedelta(days=6)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Weekly {monday.isoformat()}"

    ws.cell(row=1, column=1, value=f"Weekly Production Report - {monday.isoformat()} to {sunday.isoformat()}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(product_codes) + 2)

    headers = ["Date", "Day"] + product_codes
    _style_header(ws, 3, len(headers))

    row = 4
    for i in range(7):
        d = monday + timedelta(days=i)
        ds = d.isoformat()
        totals = get_daily_totals(ds)
        ws.cell(row=row, column=1, value=ds).border = _thin_border()
        ws.cell(row=row, column=2, value=d.strftime("%A")).border = _thin_border()
        for j, code in enumerate(product_codes):
            qty = totals.get(code, 0)
            cell = ws.cell(row=row, column=3 + j, value=qty)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(67 + j)].width = 12

    filepath = DAILY_LOGS_DIR / f"report_weekly_{monday.isoformat()}.xlsx"
    wb.save(filepath)
    return str(filepath)


def _monthly_excel_report(year: int, month: int, product_codes: list[str]) -> str:
    from calendar import monthrange
    days = monthrange(year, month)[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly {year}-{month:02d}"

    ws.cell(row=1, column=1, value=f"Monthly Production Report - {year}-{month:02d}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(product_codes) + 1)

    headers = ["Date"] + product_codes
    _style_header(ws, 3, len(headers))

    row = 4
    for d in range(1, days + 1):
        date_str = f"{year}-{month:02d}-{d:02d}"
        totals = get_daily_totals(date_str)
        has_data = any(totals.values())
        if not has_data:
            continue
        ws.cell(row=row, column=1, value=date_str).border = _thin_border()
        for j, code in enumerate(product_codes):
            qty = totals.get(code, 0)
            ws.cell(row=row, column=2 + j, value=qty).border = _thin_border()
            ws.cell(row=row, column=2 + j).alignment = Alignment(horizontal="center")
        row += 1

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 12

    filepath = DAILY_LOGS_DIR / f"report_monthly_{year}-{month:02d}.xlsx"
    wb.save(filepath)
    return str(filepath)


def generate_worker_excel(worker_name: str, year: int, month: int) -> Optional[str]:
    from calendar import monthrange
    from tools.database import get_worker_id

    wid = get_worker_id(worker_name)
    if not wid:
        return None

    products = get_all_products()
    product_codes = [p["code"] for p in products]
    days_data = get_worker_daily_breakdown(wid, year, month)

    if not days_data:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{worker_name} {year}-{month:02d}"

    num_cols = len(product_codes) + 1

    title_cell = ws.cell(row=1, column=1, value=f"{worker_name} - {year}-{month:02d} Production Report")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    headers = ["Date"] + product_codes
    _write_header_row(ws, 3, headers)

    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    absent_font = Font(color="D93025", bold=True)

    row = 4
    grand_totals = {code: 0 for code in product_codes}
    date_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    for day in days_data:
        date_cell = ws.cell(row=row, column=1, value=day["date"])
        date_cell.border = _thin_border()
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal="center")

        if day["status"] == "absent":
            reason_text = day.get("reason") or "ABSENT"
            cell = ws.cell(row=row, column=2, value=reason_text)
            cell.border = _thin_border()
            cell.fill = absent_fill
            cell.font = absent_font
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = _thin_border()
                ws.cell(row=row, column=c).fill = absent_fill
            row += 1
            continue

        for j, code in enumerate(product_codes):
            qty = day["products"].get(code, 0)
            cell = ws.cell(row=row, column=2 + j, value=qty)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
            grand_totals[code] += qty
        row += 1

    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(bold=True, size=11)
    total_cell.border = _thin_border()
    total_cell.alignment = Alignment(horizontal="center")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for j, code in enumerate(product_codes):
        cell = ws.cell(row=row, column=2 + j, value=grand_totals[code])
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.fill = total_fill

    ws.column_dimensions["A"].width = 14
    for j in range(len(product_codes)):
        ws.column_dimensions[chr(66 + j)].width = 14

    filepath = HISTORY_DIR / f"{worker_name}_{year}_{month:02d}.xlsx"
    wb.save(filepath)
    return str(filepath)


def _write_header_row(ws, row: int, headers: list[str]):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = _thin_border()
